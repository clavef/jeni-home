# cards.py v26 - 제니앱 카드값 계산기 (날짜 문제 해결 완전판)

import streamlit as st
import pandas as pd
import re
from typing import Optional
from shared import show_menu

st.set_page_config(page_title="제니앱", page_icon="💳", layout="wide")
show_menu("카드값 계산기")

st.title("💳 카드값 계산기")

# ✅ 날짜 자동 변환 헬퍼
def safe_excel_date(series):
    sample = series.dropna().iloc[0] if not series.dropna().empty else None
    if isinstance(sample, (int, float)):
        return pd.to_datetime(series, errors="coerce", unit="d", origin="1899-12-30")
    return pd.to_datetime(series, errors="coerce")

# ✅ 카드사명 정규화
def normalize_card_name(card):
    for key in [("국민", "국민카드"), ("신한", "신한카드"), ("현대", "현대카드"),
                ("하나", "하나카드"), ("로테", "로테카드"), ("삼성", "삼성카드")]:
        if key[0] in card:
            return key[1]
    return card

# ✅ 자동 카테고리 분류
def categorize(merchant: str) -> str:
    merchant = str(merchant)
    rules = [
        (r"주차장|파킹|빌딩관리단|티머니|택시|에너지|버스|도로|주유|충전|자동차|세차|오토오아시스", "교통/주유/주차"),
        (r"롯데마트|달콤N|매머드|헤이듀|한울곰탕|워커스하이|카페|커피|이디야|스타벅스|편의점|씨유|CU|GS25|세븐일레븐|emart24|올리브영|식당|음식|한솥|고기|김밥|배달", "음식점/카페/편의점"),
        (r"기프티샷|백화점|인터넷상거래|네이버페이|페이코|PAYPAL|기프티콘|쇼핑|디지털|전자|마켓|Temu|쿠팡|위메프|G마켓|11번가|인터파크|스마트스토어|번개장터", "취미/쇼핑"),
        (r"KCP|보람상조|효성에프엠에스|Microsoft|\(주\)다날\s*-\s*카카오|자동결제|관리비|통신|SKT|KT|LGU\+|렌탈|보험|납부|세금|등록금|교육비|마이데이터|고정지출", "고정지출"),
        (r"병원|치과|의원|내과|약국|정형외과", "병원/약국"),
    ]
    for pattern, category in rules:
        if re.search(pattern, merchant, re.IGNORECASE):
            return category
    return "잡비용"

# ✅ 카드사 자동 인식
def detect_card_issuer(file) -> Optional[str]:
    try:
        xls = pd.ExcelFile(file)
        def normalize(text): return str(text).replace('\n', '').replace('\r', '').replace(' ', '').strip()
        patterns = {
            "롯데카드": [{"이용일자", "이용가맹점", "업종", "이용금액"}],
            "KB국민카드": [{"이용일", "이용하신곳", "이용카드명", "국내이용금액(원)"}],
            "신한카드": [{"거래일자", "이용가맹점", "거래금액"}],
            "현대카드": [{"이용일", "이용가맹점", "이용금액"}],
            "삼성카드": [
                {"승인일자", "가맹점명", "승인금액(원)"},
                {"이용일자", "사용처/가맹점", "이용금액"},
                {"이용일자", "사용처/가맹점", "결제예정금액"},
            ],
            "하나카드": [{"거래일자", "가맹점명", "이용금액"}],
        }
        for sheet in xls.sheet_names:
            df = xls.parse(sheet, header=None)
            for i in range(min(100, len(df))):
                row = df.iloc[i]
                normed = set(normalize(cell) for cell in row if pd.notna(cell))
                for issuer, keyword_sets in patterns.items():
                    for keyword_set in keyword_sets:
                        normed_keywords = set(normalize(k) for k in keyword_set)
                        if normed_keywords.issubset(normed):
                            return issuer
        return None
    except Exception as e:
        print("[ERROR] detect_card_issuer 예외 발생:", e)
        return None

# ✅ 카드사별 파서 연결
def parse_card_file(file, issuer: str) -> Optional[pd.DataFrame]:
    parsers = {
        "롯데카드": parse_lotte,
        "KB국민카드": parse_kb,
        "신한카드": parse_shinhan,
        "현대카드": parse_hyundai,
        "하나카드": parse_hana,
        "삼성카드": parse_samsung
    }
    return parsers.get(issuer, lambda f: None)(file)

# ✅✅ 카드사별 파싱 시작
# ✅✅ 카드사별 파싱 시작

# ✅ 현대카드
def parse_hyundai(file):
    try:
        xls = pd.ExcelFile(file)
        df = xls.parse(0, skiprows=2)
        df.columns = df.columns.astype(str).str.strip()

        if not {"이용일", "이용가맹점", "이용금액"}.issubset(df.columns):
            return None

        # 병합된 문자열에서 숫자 추출
        def extract_first_number(cell):
            if isinstance(cell, str):
                numbers = re.findall(r"\d+", cell)
                if numbers:
                    return int(numbers[0])
                return None
            elif isinstance(cell, (int, float)):
                return cell
            else:
                return None

        df["이용일"] = df["이용일"].apply(extract_first_number)
        df["이용일"] = pd.to_datetime(df["이용일"], errors="coerce", unit="d", origin="1899-12-30")
        df = df[df["이용일"].notna()]
        df["이용일"] = df["이용일"].dt.strftime("%Y.%m.%d")

        df = df[["이용일", "이용가맹점", "이용금액"]]
        df.columns = ["날짜", "사용처", "금액"]
        df["카드"] = "현대카드"
        df["카테고리"] = ""

        return df[["날짜", "카드", "카테고리", "사용처", "금액"]]
    except:
        return None

# ✅ 삼성카드
def parse_samsung(file):
    try:
        def extract_excel_date(cell):
            if isinstance(cell, str):
                nums = re.findall(r"\d+", cell)
                if nums:
                    return int(nums[0])
                return None
            elif isinstance(cell, (int, float)):
                return cell
            return None

        xls = pd.ExcelFile(file)
        sheet = xls.sheet_names[0]
        raw = xls.parse(sheet, header=None)

        header_keywords_sets = [
            {"승인일자", "승인시각", "가맹점명", "승인금액(원)"},
            {"이용일자", "카드번호", "사용처/가맹점", "이용금액"},
            {"이용일자", "사용처/가맹점", "결제예정금액"},
        ]

        df = None
        for i, row in raw.iterrows():
            cells = [str(c).strip() for c in row if pd.notna(c)]
            for header_keywords in header_keywords_sets:
                if header_keywords.issubset(set(cells)):
                    df = xls.parse(sheet, skiprows=i)
                    df.columns = df.columns.astype(str).str.strip()
                    break
            if df is not None:
                break

        if df is None:
            return None

        # ✅ 승인내역 구조
        if {"승인일자", "승인시각", "가맹점명", "승인금액(원)"}.issubset(df.columns):
            df = df[["승인일자", "승인시각", "가맹점명", "승인금액(원)"]].copy()
            df["승인일자"] = df["승인일자"].apply(extract_excel_date)
            df["날짜"] = pd.to_datetime(df["승인일자"], errors="coerce", unit="d", origin="1899-12-30")
            df = df[df["날짜"].notna()]
            df["날짜"] = df["날짜"].dt.strftime("%Y.%m.%d")
            df["사용처"] = df["가맹점명"]
            df["금액"] = df["승인금액(원)"].astype(str).str.replace(",", "").astype(float)
            df["카드"] = "삼성카드"
            df["카테고리"] = ""
            return df[["날짜", "카드", "카테고리", "사용처", "금액"]]

        # ✅ 리볼빙 구조
        if {"이용일자", "카드번호", "사용처/가맹점", "이용금액"}.issubset(df.columns):
            df = df[["이용일자", "사용처/가맹점", "이용금액"]].copy()
            df.columns = ["날짜", "사용처", "금액"]
            df["날짜"] = df["날짜"].apply(extract_excel_date)
            df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce", unit="d", origin="1899-12-30")
            df = df[df["날짜"].notna()]
            df["날짜"] = df["날짜"].dt.strftime("%Y.%m.%d")
            df["금액"] = df["금액"].astype(str).str.replace(",", "").astype(float)
            df["카드"] = "삼성카드"
            df["카테고리"] = ""
            return df[["날짜", "카드", "카테고리", "사용처", "금액"]]

        # ✅ 연회비 구조
        if {"이용일자", "사용처/가맹점", "결제예정금액"}.issubset(df.columns):
            df = df[["이용일자", "사용처/가맹점", "결제예정금액"]].copy()
            df.columns = ["날짜", "사용처", "금액"]
            df["날짜"] = df["날짜"].apply(extract_excel_date)
            df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce", unit="d", origin="1899-12-30")
            df = df[df["날짜"].notna()]
            df["날짜"] = df["날짜"].dt.strftime("%Y.%m.%d")
            df["금액"] = df["금액"].astype(str).str.replace(",", "").astype(float)
            df["카드"] = "삼성카드"
            df["카테고리"] = ""
            return df[["날짜", "카드", "카테고리", "사용처", "금액"]]

        return None
    except:
        return None

# ✅ 롯데카드
def parse_lotte(file):
    try:
        xls = pd.ExcelFile(file)
        sheet = xls.sheet_names[0]
        raw = xls.parse(sheet, header=None)
        header_keywords = {"이용일자", "이용가맹점", "업종", "이용금액"}

        for i, row in raw.iterrows():
            cells = [str(c).strip() for c in row if pd.notna(c)]
            if header_keywords.issubset(set(cells)):
                df = xls.parse(sheet, skiprows=i)
                break
        else:
            return None

        df.columns = df.columns.str.strip()
        if "취소여부" in df.columns:
            df = df[df["취소여부"].astype(str).str.upper() != "Y"]

        df = df[["이용일자", "이용가맹점", "업종", "이용금액"]].copy()
        df.columns = ["날짜", "사용처", "카테고리", "금액"]
        df["카드"] = "롯데카드"
        return df[["날짜", "카드", "카테고리", "사용처", "금액"]]
    except:
        return None

# ✅ 국민카드
def parse_kb(file):
    try:
        xls = pd.ExcelFile(file)
        df = xls.parse(xls.sheet_names[0], skiprows=6)
        if "상태" in df.columns:
            df = df[~df["상태"].astype(str).str.contains("승인취소|취소전표", na=False)]

        df = df[["이용일", "이용하신곳", "이용카드명", "국내이용금액\n(원)", "결제방법"]]
        df.columns = ["날짜", "사용처", "카드", "금액", "결제방법"]
        df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.strftime("%Y.%m.%d")
        df["금액"] = df["금액"].astype(str).str.replace(",", "").astype(int)

        def adjust(row):
            method = str(row["결제방법"])
            if method != "일시불" and any(char.isdigit() for char in method):
                return round(row["금액"] / int(''.join(filter(str.isdigit, method))))
            return row["금액"]
        df["금액"] = df.apply(adjust, axis=1)

        df["카테고리"] = ""
        return df[["날짜", "카드", "카테고리", "사용처", "금액"]]
    except:
        return None

# ✅ 신한카드
def parse_shinhan(file):
    try:
        df = pd.ExcelFile(file).parse(0, skiprows=2)
        df = df[["거래일자", "이용가맹점", "결제 금액"]]
        df.columns = ["날짜", "사용처", "금액"]
        df["금액"] = pd.to_numeric(df["금액"], errors="coerce")
        df["카드"] = "신한카드"
        df["카테고리"] = ""
        return df[["날짜", "카드", "카테고리", "사용처", "금액"]]
    except:
        return None

# ✅ 하나카드
def parse_hana(file):
    try:
        df = pd.ExcelFile(file).parse(0, skiprows=28)
        df.columns = df.columns.astype(str).str.replace('\n', '').str.replace(' ', '').str.strip()
        if not {"거래일자", "가맹점명", "이용금액"}.issubset(df.columns):
            return None

        df = df[["거래일자", "가맹점명", "이용금액"]]
        df.columns = ["날짜", "사용처", "금액"]
        df["카드"] = "하나카드"
        df["카테고리"] = ""

        # 날짜를 to_datetime으로 정리 (엑셀 날짜 형식 포함)
        df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
        df = df[df["날짜"].notna()]  # 유효한 날짜만 남김
        df["날짜"] = df["날짜"].dt.strftime("%Y.%m.%d")

        return df[["날짜", "카드", "카테고리", "사용처", "금액"]]
    except:
        return None

# ✅✅ 카드사별 파싱 종료
# ✅✅ 카드사별 파싱 종료

# ✅ 파일 업로드
uploaded_files = st.file_uploader(
    "카드사별 이용 내역 파일 업로드 (여러 개 가능)",
    type=["xlsx"],
    accept_multiple_files=True
)

# ✅ 처리 시작
if uploaded_files:
    all_records = []
    for file in uploaded_files:
        st.markdown(f"---\n### 📂 {file.name}")
        card_issuer = detect_card_issuer(file)
        if not card_issuer:
            st.warning(f"❌ 카드사 인식 실패: {file.name}")
            continue
        df = parse_card_file(file, card_issuer)
        if df is not None:
            all_records.append(df)
            st.success(f"✅ {card_issuer} 내역 처리 완료: {len(df)}건")
        else:
            st.warning(f"⚠️ {card_issuer} 내역 파싱 실패")

    if all_records:
        final_df = pd.concat(all_records, ignore_index=True)
        final_df["카드"] = final_df["카드"].apply(normalize_card_name)
        final_df["카테고리"] = final_df["사용처"].apply(categorize)
        final_df["금액"] = final_df["금액"].apply(lambda x: float(str(x).replace(",", "")))
        final_df = final_df.sort_values(by=["카드", "카테고리", "날짜"]).reset_index(drop=True)

        st.subheader("📋 통합 카드 사용 내역")
        st.dataframe(final_df, use_container_width=True)

        @st.cache_data
        def to_excel(df):
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.page import PageMargins
            from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
            from openpyxl.chart import BarChart, Reference

            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = '카드내역'

            card_list = ["국민카드", "현대카드", "롯데카드", "삼성카드", "하나카드", "신한카드"]
            color_map_card = {
                "국민카드": "FBE2D5", "현대카드": "DDEBF7", "롯데카드": "CCCCFF",
                "삼성카드": "E2EFDA", "하나카드": "FFF2CC", "신한카드": "DDD9C4",
            }

            color_map_category = {
                "교통/주유/주차": "CCFFCC", "병원/약국": "FFCC99",
                "취미/쇼핑": "FFF2CC", "음식점/카페/편의점": "FFCCCC",
                "고정지출": "C6E0B4", "잡비용": "E7E6E6",
            }

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # 헤더
            ws.append(df.columns.tolist())
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="000000")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

            # 열 너비
            for i, width in enumerate([11, 11, 20, 40, 15]):
                ws.column_dimensions[chr(65 + i)].width = width
            ws.column_dimensions['F'].width = 3
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 3
            ws.sheet_view.showGridLines = False

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                card = row[1].value
                category = row[2].value
                card_color = color_map_card.get(card, "E7E6E6")
                cat_color = color_map_category.get(category, "E7E6E6")
                for idx, cell in enumerate(row):
                    cell.border = thin_border
                    if idx == 4:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                row[0].fill = row[1].fill = PatternFill("solid", fgColor=card_color)
                row[2].fill = PatternFill("solid", fgColor=cat_color)

            # 카테고리별 통계
            ws["G1"] = "카테고리"
            ws["H1"] = "금액"
            ws["G1"].fill = ws["H1"].fill = PatternFill("solid", fgColor="000000")
            ws["G1"].font = ws["H1"].font = Font(color="FFFFFF", bold=True)
            ws["G1"].alignment = ws["H1"].alignment = Alignment(horizontal="center", vertical="center")

            stats = df.groupby("카테고리")["금액"].sum()
            row_idx = 2
            for cat, amount in stats.items():
                ws[f"G{row_idx}"] = cat
                ws[f"H{row_idx}"] = int(amount)
                ws[f"H{row_idx}"].number_format = '#,##0'
                ws[f"H{row_idx}"].font = Font(bold=False)
                cat_color = color_map_category.get(cat, "E7E6E6")
                ws[f"G{row_idx}"].fill = PatternFill("solid", fgColor=cat_color)
                ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border
                row_idx += 1

            ws[f"G{row_idx}"] = "합계"
            ws[f"H{row_idx}"] = int(stats.sum())
            ws[f"G{row_idx}"].fill = ws[f"H{row_idx}"].fill = PatternFill("solid", fgColor="000000")
            ws[f"G{row_idx}"].font = Font(color="FFFFFF", bold=True)
            ws[f"H{row_idx}"].font = Font(color="FFFFFF", bold=True)
            ws[f"H{row_idx}"].number_format = '#,##0'
            ws[f"G{row_idx}"].alignment = ws[f"H{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border

            cat_rows = row_idx - 1

            # 카드사별 통계
            ws["G10"] = "카드사"
            ws["H10"] = "금액"
            ws["G10"].fill = ws["H10"].fill = PatternFill("solid", fgColor="000000")
            ws["G10"].font = ws["H10"].font = Font(color="FFFFFF", bold=True)
            ws["G10"].alignment = ws["H10"].alignment = Alignment(horizontal="center", vertical="center")

            stats2 = df.groupby("카드")["금액"].sum().reindex(card_list, fill_value=0)
            row_idx = 11
            for card, amount in stats2.items():
                ws[f"G{row_idx}"] = card
                ws[f"H{row_idx}"] = int(amount)
                ws[f"H{row_idx}"].number_format = '#,##0'
                ws[f"H{row_idx}"].font = Font(bold=False)
                card_color = color_map_card.get(card, "E7E6E6")
                ws[f"G{row_idx}"].fill = PatternFill("solid", fgColor=card_color)
                ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border
                row_idx += 1

            ws[f"G{row_idx}"] = "합계"
            ws[f"H{row_idx}"] = int(stats2.sum())
            ws[f"G{row_idx}"].fill = ws[f"H{row_idx}"].fill = PatternFill("solid", fgColor="000000")
            ws[f"G{row_idx}"].font = Font(color="FFFFFF", bold=True)
            ws[f"H{row_idx}"].font = Font(color="FFFFFF", bold=True)
            ws[f"H{row_idx}"].number_format = '#,##0'
            ws[f"G{row_idx}"].alignment = ws[f"H{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"G{row_idx}"].border = ws[f"H{row_idx}"].border = thin_border

            # 막대형 차트 1: 카테고리
            bar1 = BarChart()
            bar1.type = "bar"
            bar1.style = 10
            bar1.y_axis.majorGridlines = None
            bar1.legend = None
            bar1.title = None
            bar1.height = 6
            bar1.width = 5
            data1 = Reference(ws, min_col=8, min_row=1, max_row=1 + cat_rows)
            cats1 = Reference(ws, min_col=7, min_row=2, max_row=1 + cat_rows)
            bar1.add_data(data1, titles_from_data=True)
            bar1.set_categories(cats1)
            bar1.x_axis.delete = True
            ws.add_chart(bar1, "J1")

            # 막대형 차트 2: 카드사
            bar2 = BarChart()
            bar2.type = "bar"
            bar2.style = 10
            bar2.y_axis.majorGridlines = None
            bar2.legend = None
            bar2.title = None
            bar2.height = 6
            bar2.width = 5
            data2 = Reference(ws, min_col=8, min_row=10, max_row=10 + len(card_list))
            cats2 = Reference(ws, min_col=7, min_row=11, max_row=10 + len(card_list))
            bar2.add_data(data2, titles_from_data=True)
            bar2.set_categories(cats2)
            bar2.x_axis.delete = True
            ws.add_chart(bar2, "J14")

            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
            ws.sheet_properties = WorksheetProperties(pageSetUpPr=PageSetupProperties(fitToPage=True))
            wb.save(output)
            return output.getvalue()

        st.download_button(
            label="📅 엑셀파일 다운로드",
            data=to_excel(final_df),
            file_name="카드값_통합내역.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
